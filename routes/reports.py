"""
PDF Report generation routes
"""
from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from app.database import get_db
from app.models.client import Client
from app.models.scenario import Scenario
from models.scenario_cashflow import ScenarioCashflow
from utils.pdf_builder import build_pdf
from utils.logging_decorators import log_calculation
from typing import Dict, Any, List
import tempfile
import os
from datetime import datetime
import uuid

router = APIRouter(prefix="/api/v1/reports", tags=["reports"])

@log_calculation
def prepare_report_context(client: Client, scenario: Scenario, db: Session) -> Dict[str, Any]:
    """
    Prepare context data for PDF report generation
    """
    
    # Basic client and scenario info
    context = {
        'client': client,
        'scenario': scenario,
        'retirement_age': scenario.parameters.get('retirement_age', 67),
        'life_expectancy': scenario.parameters.get('life_expectancy', 85),
        'indexation_rate': scenario.parameters.get('indexation_rate', 0.02)
    }
    
    # Pension calculation details
    pension_calc = scenario.parameters.get('pension_calculation', {})
    context['pension_calculation'] = {
        'total_capital': pension_calc.get('total_capital', 0),
        'reservation_impact': pension_calc.get('reservation_impact', 0),
        'effective_capital': pension_calc.get('effective_capital', 0),
        'years_in_retirement': context['life_expectancy'] - context['retirement_age'],
        'annual_pension': pension_calc.get('annual_pension', 0),
        'monthly_pension': pension_calc.get('annual_pension', 0) / 12
    }
    
    # Processed grants
    grants = scenario.parameters.get('processed_grants', [])
    context['processed_grants'] = grants
    
    # Tax parameters
    context['tax_parameters'] = {
        'brackets': [
            {'min': 0, 'max': 75960, 'rate': 0.10},
            {'min': 75960, 'max': 108960, 'rate': 0.14},
            {'min': 108960, 'max': 174960, 'rate': 0.20},
            {'min': 174960, 'max': 243120, 'rate': 0.31},
            {'min': 243120, 'max': 504360, 'rate': 0.35},
            {'min': 504360, 'max': 663240, 'rate': 0.47},
            {'min': 663240, 'max': None, 'rate': 0.50}
        ]
    }
    
    return context

@log_calculation
def prepare_cashflow_data(scenario: Scenario, db: Session) -> List[Dict[str, Any]]:
    """
    Prepare cashflow data for PDF report
    """
    
    # Get cashflow records
    cashflow_records = db.query(ScenarioCashflow).filter(
        ScenarioCashflow.scenario_id == scenario.id
    ).order_by(ScenarioCashflow.year).all()
    
    if not cashflow_records:
        return []
    
    # Convert to list of dictionaries
    cashflow_data = []
    for record in cashflow_records:
        cashflow_data.append({
            'year': record.year,
            'gross_income': float(record.gross_income or 0),
            'pension_income': float(record.pension_income or 0),
            'grant_income': float(record.grant_income or 0),
            'other_income': float(record.other_income or 0),
            'tax': float(record.tax or 0),
            'net_income': float(record.net_income or 0)
        })
    
    return cashflow_data

@router.post("/scenario/{scenario_id}/pdf")
async def generate_scenario_pdf_report(
    scenario_id: int,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db)
):
    """
    Generate PDF report for a scenario
    """
    
    # Get scenario with client
    scenario = db.query(Scenario).filter(Scenario.id == scenario_id).first()
    if not scenario:
        raise HTTPException(status_code=404, detail="Scenario not found")
    
    # Get client
    client = db.query(Client).filter(Client.id == scenario.client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    try:
        # Prepare report context
        context = prepare_report_context(client, scenario, db)
        
        # Prepare cashflow data
        cashflow_data = prepare_cashflow_data(scenario, db)
        
        if not cashflow_data:
            raise HTTPException(
                status_code=400, 
                detail="No cashflow data available for this scenario. Please run the scenario first."
            )
        
        # Generate unique filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"retirement_report_{client.id}_{scenario_id}_{timestamp}.pdf"
        
        # Create temporary file
        temp_dir = tempfile.gettempdir()
        file_path = os.path.join(temp_dir, filename)
        
        # Build PDF
        build_pdf(context, cashflow_data, file_path)
        
        # Schedule cleanup after response
        background_tasks.add_task(cleanup_temp_file, file_path)
        
        # Return file response
        return FileResponse(
            path=file_path,
            filename=filename,
            media_type='application/pdf',
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate PDF: {str(e)}")

@router.post("/client/{client_id}/comprehensive-report")
async def generate_comprehensive_client_report(
    client_id: int,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db)
):
    """
    Generate comprehensive PDF report for all client scenarios
    """
    
    # Get client
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Get all completed scenarios for client
    scenarios = db.query(Scenario).filter(
        Scenario.client_id == client_id,
        Scenario.status == 'completed'
    ).all()
    
    if not scenarios:
        raise HTTPException(
            status_code=400,
            detail="No completed scenarios found for this client"
        )
    
    try:
        # Use the first scenario for main report (or could combine multiple)
        main_scenario = scenarios[0]
        
        # Prepare report context
        context = prepare_report_context(client, main_scenario, db)
        
        # Add comparison data if multiple scenarios
        if len(scenarios) > 1:
            context['scenario_comparison'] = []
            for scenario in scenarios:
                cashflow = prepare_cashflow_data(scenario, db)
                if cashflow:
                    total_net = sum(entry['net_income'] for entry in cashflow)
                    context['scenario_comparison'].append({
                        'scenario_name': scenario.name,
                        'total_net_income': total_net,
                        'years_covered': len(cashflow)
                    })
        
        # Prepare cashflow data
        cashflow_data = prepare_cashflow_data(main_scenario, db)
        
        if not cashflow_data:
            raise HTTPException(
                status_code=400,
                detail="No cashflow data available. Please run scenarios first."
            )
        
        # Generate unique filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"comprehensive_report_{client_id}_{timestamp}.pdf"
        
        # Create temporary file
        temp_dir = tempfile.gettempdir()
        file_path = os.path.join(temp_dir, filename)
        
        # Build PDF
        build_pdf(context, cashflow_data, file_path)
        
        # Schedule cleanup after response
        background_tasks.add_task(cleanup_temp_file, file_path)
        
        # Return file response
        return FileResponse(
            path=file_path,
            filename=filename,
            media_type='application/pdf',
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate comprehensive report: {str(e)}")

@router.get("/scenario/{scenario_id}/preview")
async def preview_scenario_report_data(
    scenario_id: int,
    db: Session = Depends(get_db)
):
    """
    Preview report data without generating PDF (for debugging/testing)
    """
    
    # Get scenario with client
    scenario = db.query(Scenario).filter(Scenario.id == scenario_id).first()
    if not scenario:
        raise HTTPException(status_code=404, detail="Scenario not found")
    
    # Get client
    client = db.query(Client).filter(Client.id == scenario.client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    try:
        # Prepare report context
        context = prepare_report_context(client, scenario, db)
        
        # Prepare cashflow data
        cashflow_data = prepare_cashflow_data(scenario, db)
        
        # Return preview data
        return {
            "context": {
                "client_name": getattr(client, 'full_name', 'Unknown'),
                "scenario_name": scenario.name,
                "retirement_age": context['retirement_age'],
                "life_expectancy": context['life_expectancy'],
                "pension_calculation": context['pension_calculation'],
                "grants_count": len(context['processed_grants']),
                "tax_brackets_count": len(context['tax_parameters']['brackets'])
            },
            "cashflow_summary": {
                "years_count": len(cashflow_data),
                "total_gross_income": sum(entry['gross_income'] for entry in cashflow_data),
                "total_net_income": sum(entry['net_income'] for entry in cashflow_data),
                "total_tax": sum(entry['tax'] for entry in cashflow_data),
                "first_year": min(entry['year'] for entry in cashflow_data) if cashflow_data else None,
                "last_year": max(entry['year'] for entry in cashflow_data) if cashflow_data else None
            },
            "sample_cashflow": cashflow_data[:5] if cashflow_data else []  # First 5 years
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to preview report: {str(e)}")

@router.post("/clients/{client_id}/reports/pdf")
async def generate_client_pdf_report(
    client_id: int,
    request: dict,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db)
):
    """
    Generate PDF report for client - Frontend compatibility endpoint
    """
    
    # Get client
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Get scenario ID from request
    scenario_id = request.get('scenario_id', 1)
    
    # Get scenario
    scenario = db.query(Scenario).filter(
        Scenario.id == scenario_id,
        Scenario.client_id == client_id
    ).first()
    
    if not scenario:
        # Get any available scenario for the client
        scenario = db.query(Scenario).filter(
            Scenario.client_id == client_id
        ).first()
        
        if not scenario:
            # Create a default scenario if none exists
            scenario = Scenario(
                client_id=client_id,
                name="דוח ברירת מחדל",
                parameters={},
                status='draft'
            )
            db.add(scenario)
            db.commit()
            db.refresh(scenario)
    
    try:
        # Prepare report context
        context = prepare_report_context(client, scenario, db)
        
        # Prepare cashflow data
        cashflow_data = prepare_cashflow_data(scenario, db)
        
        # If no cashflow data, create mock data for demo
        if not cashflow_data:
            current_year = datetime.now().year
            cashflow_data = []
            for i in range(20):  # 20 years of mock data
                year = current_year + i
                cashflow_data.append({
                    'year': year,
                    'gross_income': 150000 + (i * 2000),  # Growing income
                    'pension_income': 80000 if i >= 5 else 0,  # Pension starts after 5 years
                    'grant_income': 50000 if i < 10 else 0,  # Grants for first 10 years
                    'other_income': 20000,
                    'tax': 45000 + (i * 500),  # Growing tax
                    'net_income': 125000 + (i * 1500)  # Growing net income
                })
        
        # Generate unique filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"retirement_report_{client_id}_{timestamp}.pdf"
        
        # Create temporary file
        temp_dir = tempfile.gettempdir()
        file_path = os.path.join(temp_dir, filename)
        
        # Build PDF
        build_pdf(context, cashflow_data, file_path)
        
        # Schedule cleanup after response
        background_tasks.add_task(cleanup_temp_file, file_path)
        
        # Return file response
        return FileResponse(
            path=file_path,
            filename=filename,
            media_type='application/pdf',
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate PDF: {str(e)}")

@router.post("/clients/{client_id}/reports/excel")
async def generate_client_excel_report(
    client_id: int,
    request: dict,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db)
):
    """
    Generate Excel report for client - Frontend compatibility endpoint
    """
    
    # Get client
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    try:
        # Get report data from request
        report_data = request.get('report_data', {})
        
        # Create Excel content using openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "דוח פרישה"
        
        # Header style
        header_font = Font(bold=True, size=14)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Client info section
        ws['A1'] = 'דוח פרישה'
        ws['A1'].font = Font(bold=True, size=16)
        ws['A3'] = 'פרטי לקוח:'
        ws['A3'].font = header_font
        
        client_info = report_data.get('client_info', {})
        ws['A4'] = f"שם: {client_info.get('name', 'לא צוין')}"
        ws['A5'] = f"ת.ז.: {client_info.get('id_number', 'לא צוין')}"
        
        # Grants summary section
        ws['A7'] = 'סיכום מענקים:'
        ws['A7'].font = header_font
        
        grants_summary = report_data.get('grants_summary', {})
        ws['A8'] = f"סך כל מענקים: ₪{grants_summary.get('total_grants', 0):,}"
        ws['A9'] = f"סך פטור ממס: ₪{grants_summary.get('total_exempt', 0):,}"
        ws['A10'] = f"סך חייב במס: ₪{grants_summary.get('total_taxable', 0):,}"
        ws['A11'] = f"מס משוער: ₪{grants_summary.get('estimated_tax', 0):,}"
        
        # Yearly totals section
        ws['A13'] = 'סיכום שנתי:'
        ws['A13'].font = header_font
        
        yearly_totals = report_data.get('yearly_totals', {})
        ws['A14'] = f"סך הכנסות: ₪{yearly_totals.get('total_income', 0):,}"
        ws['A15'] = f"סך מסים: ₪{yearly_totals.get('total_tax', 0):,}"
        ws['A16'] = f"הכנסה נטו: ₪{yearly_totals.get('net_income', 0):,}"
        
        # Cashflow projection section
        ws['A18'] = 'תחזית תזרים מזומנים:'
        ws['A18'].font = header_font
        
        # Headers for cashflow table
        ws['A19'] = 'תאריך'
        ws['B19'] = 'סכום'
        ws['C19'] = 'מקור'
        
        for cell in ['A19', 'B19', 'C19']:
            ws[cell].font = header_font
            ws[cell].fill = header_fill
        
        # Cashflow data
        cashflow_projection = report_data.get('cashflow_projection', [])
        for i, entry in enumerate(cashflow_projection, start=20):
            ws[f'A{i}'] = entry.get('date', '')
            ws[f'B{i}'] = f"₪{entry.get('amount', 0):,}"
            ws[f'C{i}'] = entry.get('source', '')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Generate unique filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"retirement_report_{client_id}_{timestamp}.xlsx"
        
        # Create temporary file
        temp_dir = tempfile.gettempdir()
        file_path = os.path.join(temp_dir, filename)
        
        # Save Excel file
        wb.save(file_path)
        
        # Schedule cleanup after response
        background_tasks.add_task(cleanup_temp_file, file_path)
        
        # Return file response
        return FileResponse(
            path=file_path,
            filename=filename,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate Excel: {str(e)}")

def cleanup_temp_file(file_path: str):
    """
    Clean up temporary PDF file
    """
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
    except Exception:
        pass  # Ignore cleanup errors
